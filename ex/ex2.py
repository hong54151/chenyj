from argopy import DataFetcher as ArgoDataFetcher
import xlrd
import xlwt
import openpyxl
argo_loader = ArgoDataFetcher()
ds = argo_loader.region([40, 100, 0, 30, 0, 10, '2010-01-01', '2010-01-31']).to_xarray() # 读取区域包含的多个浮标多期的数据
pr=ds['PLATFORM_NUMBER'].values
dss = argo_loader.profile(pr[1], 2).to_xarray() # 读取一个浮标一期的数据，2为浮标周期数
dss_profiles = dss.argo.point2profile()
temp = dss_profiles['TEMP'].values
pres = dss_profiles['PRES'].values
#path='D:\wxd\数据.xls'
#data=xlwt.Workbook()
#data1=data.add_sheet('sheet1')
#data1.write(1,2,'hello')
#data1.save(r"D:\wxd\shuju.xls")
data = openpyxl.Workbook() # 新建工作簿
data.create_sheet('Sheet1') # 添加页
table = data.active # 获得当前活跃的工作页，默认为第一个工作页
table.cell(1,1,'Test') # 行，列，值 这里是从1开始计数的
#print(temp.shape[1])
n=temp.shape[1]
print(n)
print(temp)
#for i in range(1,n):
#    table.cell(i,2,temp[i])

data.save('excel_test.xlsx')
#xx=xlrd.open_workbook(r"D:\wxd\shuju.xls")
#xs=xx.sheet_by_name('Sheet1')
#print(xs.row_values(1))
#G=xs.col_values(2)
#print(G)